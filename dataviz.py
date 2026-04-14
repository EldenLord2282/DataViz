#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════╗
║           DataViz Terminal Dashboard  v2.1                       ║
║   Python · Matplotlib · Pandas · CSV Import · Excel Export       ║
╚══════════════════════════════════════════════════════════════════╝
"""

import sys
import os
import io
import datetime

# ── dependency check ────────────────────────────────────────────────
def check_deps():
    missing = []
    for pkg in ['matplotlib', 'pandas', 'numpy', 'openpyxl']:
        try:
            __import__(pkg)
        except ImportError:
            missing.append(pkg)
    if missing:
        print(f"\n[ERROR] Missing packages: {', '.join(missing)}")
        print(f"  Run:  pip install {' '.join(missing)}\n")
        sys.exit(1)

check_deps()

import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ── colour palette ───────────────────────────────────────────────────
BLUE   = '#58a6ff'
GREEN  = '#3fb950'
PURPLE = '#d2a8ff'
ORANGE = '#ffa657'
RED    = '#f78166'
CYAN   = '#39d353'
GOLD   = '#e3b341'
PINK   = '#f0883e'
COLORS = [BLUE, GREEN, PURPLE, ORANGE, RED, CYAN, GOLD, PINK]

BG         = '#0d1117'
BG2        = '#161b22'
BG3        = '#1c2230'
BORDER_CLR = '#30363d'
TEXT       = '#e6edf3'
TEXT2      = '#8b949e'
TEXT3      = '#6e7681'

# ╔══════════════════════════════════════════════════════════════════╗
# ║  TERMINAL HELPERS                                               ║
# ╚══════════════════════════════════════════════════════════════════╝

def cls():
    os.system('cls' if os.name == 'nt' else 'clear')

def cprint(text, color='', bold=False, end='\n'):
    codes = {
        'blue':   '\033[94m', 'green':  '\033[92m', 'yellow': '\033[93m',
        'red':    '\033[91m', 'cyan':   '\033[96m', 'purple': '\033[95m',
        'white':  '\033[97m', 'gray':   '\033[90m', '':       ''
    }
    print(f"{'[1m' if bold else ''}{codes.get(color,'')}{text}\033[0m", end=end)

def divider(char='─', width=66, color='gray'):
    cprint(char * width, color)

def prompt(msg, color='cyan'):
    codes = {'cyan': '\033[96m', 'green': '\033[92m', 'yellow': '\033[93m'}
    return input(f"{codes.get(color,'')}{msg}\033[0m")

def header():
    cls()
    print()
    cprint('  ╔══════════════════════════════════════════════════════════════╗', 'blue')
    cprint('  ║', 'blue', end='')
    print('\033[1m\033[96m  DataViz Terminal Dashboard  v2.1\033[0m\033[94m                       ║')
    cprint('  ║   Python · Matplotlib · Pandas · CSV · Excel Export          ║', 'blue')
    cprint('  ╚══════════════════════════════════════════════════════════════╝', 'blue')
    print()

def print_df(df, max_rows=8):
    divider()
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', 100)
    pd.set_option('display.float_format', '{:.2f}'.format)
    cprint(df.head(max_rows).to_string(index=False), 'white')
    divider()

# ╔══════════════════════════════════════════════════════════════════╗
# ║  MATPLOTLIB STYLE                                               ║
# ╚══════════════════════════════════════════════════════════════════╝

def setup_style():
    plt.rcParams.update({
        'figure.facecolor':  BG,  'axes.facecolor':    BG2,
        'axes.edgecolor':    BORDER_CLR, 'axes.labelcolor': TEXT2,
        'axes.titlecolor':   TEXT, 'axes.titlesize':   11,
        'axes.titleweight':  'bold', 'axes.titlepad':  12,
        'axes.grid':         True, 'grid.color':       BORDER_CLR,
        'grid.linewidth':    0.5,  'grid.alpha':       0.6,
        'xtick.color':       TEXT3, 'ytick.color':     TEXT3,
        'xtick.labelsize':   9,    'ytick.labelsize':  9,
        'text.color':        TEXT,  'font.family':     'monospace',
        'legend.facecolor':  BG3,  'legend.edgecolor': BORDER_CLR,
        'legend.labelcolor': TEXT2, 'legend.fontsize':  9,
        'figure.titlesize':  13,   'figure.titleweight': 'bold',
    })

def _ax_style(ax):
    ax.set_facecolor(BG2)
    ax.tick_params(colors=TEXT3)
    ax.grid(True, color=BORDER_CLR, alpha=0.5)
    for spine in ax.spines.values():
        spine.set_edgecolor(BORDER_CLR)

def add_watermark(fig, label):
    fig.text(0.99, 0.01, f'dataviz.py v2.1 · {label}',
             ha='right', va='bottom', fontsize=8,
             color=TEXT3, style='italic', family='monospace')

# ╔══════════════════════════════════════════════════════════════════╗
# ║  CHART BUILDERS  (figures are built but NOT shown yet)          ║
# ╚══════════════════════════════════════════════════════════════════╝

def make_line_fig(months, series, title_suffix=''):
    fig, ax = plt.subplots(figsize=(10, 5.5), facecolor=BG)
    _ax_style(ax)
    ax.set_title(f'plt.plot()  ·  {title_suffix}', color=BLUE, pad=10)
    for i, (name, vals) in enumerate(series.items()):
        c = COLORS[i % len(COLORS)]
        ax.plot(months, vals, color=c, linewidth=2.2,
                marker='o', markersize=5, label=name)
        ax.fill_between(months, vals, alpha=0.08, color=c)
    ax.legend()
    fig.tight_layout()
    plt.close(fig)          # keep in memory, don't display yet
    return fig

def make_bar_fig(categories, values, ylabel, title_suffix=''):
    fig, ax = plt.subplots(figsize=(10, 5.5), facecolor=BG)
    _ax_style(ax)
    ax.set_title(f'plt.bar()  ·  {title_suffix}', color=GREEN, pad=10)
    bar_colors = (COLORS * 4)[:len(categories)]
    bars = ax.bar(categories, values, color=bar_colors,
                  edgecolor=BG, linewidth=0.8, width=0.6)
    for bar, val in zip(bars, values):
        ax.text(bar.get_x() + bar.get_width() / 2,
                bar.get_height() * 1.01, f'{val:.1f}',
                ha='center', va='bottom', fontsize=8,
                color=TEXT2, fontfamily='monospace')
    ax.set_ylabel(ylabel, color=TEXT2)
    ax.tick_params(axis='x', rotation=20)
    ax.grid(True, color=BORDER_CLR, alpha=0.5, axis='y')
    fig.tight_layout()
    plt.close(fig)
    return fig

def make_pie_fig(labels, values, title_suffix=''):
    fig, ax = plt.subplots(figsize=(8, 6.5), facecolor=BG)
    ax.set_facecolor(BG2)
    ax.set_title(f'plt.pie()  ·  {title_suffix}', color=PURPLE, pad=10)
    pie_colors = (COLORS * 4)[:len(labels)]
    _, _, autotexts = ax.pie(
        values, labels=labels, autopct='%1.1f%%',
        colors=pie_colors, startangle=90,
        wedgeprops=dict(width=0.55, edgecolor=BG, linewidth=2),
        textprops=dict(color=TEXT2, fontsize=8, fontfamily='monospace')
    )
    for at in autotexts:
        at.set_color(BG)
        at.set_fontweight('bold')
    fig.tight_layout()
    plt.close(fig)
    return fig

def make_scatter_fig(x, y, xlabel, ylabel, title_suffix=''):
    fig, ax = plt.subplots(figsize=(10, 5.5), facecolor=BG)
    _ax_style(ax)
    ax.set_title(f'plt.scatter()  ·  {title_suffix}', color=ORANGE, pad=10)
    ax.scatter(x, y, color=ORANGE, alpha=0.78,
               edgecolors=BG2, linewidth=0.8, s=65, zorder=3)
    if len(x) > 2:
        try:
            z  = np.polyfit(x, y, 1)
            xr = np.linspace(min(x), max(x), 100)
            ax.plot(xr, np.poly1d(z)(xr), color=BLUE,
                    linewidth=1.5, linestyle='--', alpha=0.7)
        except Exception:
            pass
    ax.set_xlabel(xlabel, color=TEXT2)
    ax.set_ylabel(ylabel, color=TEXT2)
    fig.tight_layout()
    plt.close(fig)
    return fig

def make_barh_fig(categories, values, xlabel, title_suffix=''):
    fig, ax = plt.subplots(figsize=(10, 5.5), facecolor=BG)
    _ax_style(ax)
    ax.set_title(f'plt.barh()  ·  {title_suffix}', color=CYAN, pad=10)
    pairs   = sorted(zip(values, categories))
    sv, sc  = zip(*pairs)
    ax.barh(list(sc), list(sv), color=CYAN, alpha=0.82,
            edgecolor=BG, height=0.6)
    ax.set_xlabel(xlabel, color=TEXT2)
    ax.tick_params(axis='y', labelsize=8)
    ax.grid(True, color=BORDER_CLR, alpha=0.5, axis='x')
    fig.tight_layout()
    plt.close(fig)
    return fig

# ╔══════════════════════════════════════════════════════════════════╗
# ║  CHART PICKER  — the single place where charts are displayed    ║
# ╚══════════════════════════════════════════════════════════════════╝
#
#  all_figs : list of  (chart_type_str, display_label, fig_object)
#  Only plt.show(block=False) calls in THIS function — nowhere else.

CHART_TAG_COLOR = {
    'line':    '\033[94m',   # blue
    'bar':     '\033[92m',   # green
    'pie':     '\033[95m',   # purple
    'barh':    '\033[96m',   # cyan
    'scatter': '\033[93m',   # yellow
}

def show_fig(fig, label):
    """Reopen a figure in a new window (works even after plt.close)."""
    add_watermark(fig, label)
    dummy = plt.figure()          # create a fresh manager
    fig.set_visible(True)
    # Swap the canvas: attach our fig to the new manager's canvas
    new_manager = dummy.canvas.manager
    new_manager.canvas.figure = fig
    fig.set_canvas(new_manager.canvas)
    dummy.canvas.figure = plt.figure()  # detach dummy
    plt.close(dummy)
    fig._axstack if hasattr(fig, '_axstack') else None
    # Simplest cross-platform approach: re-draw via a new figure copy
    _show_via_savefig(fig, label)

def _show_via_savefig(fig, label):
    """
    Safe cross-platform display: save to a temp PNG, open in a new plt figure.
    This avoids TkAgg manager-swapping quirks on Windows.
    """
    import tempfile
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=120, bbox_inches='tight',
                facecolor=fig.get_facecolor())
    buf.seek(0)

    # Read back and display in a fresh figure
    img = plt.imread(buf)
    disp = plt.figure(figsize=(fig.get_figwidth(), fig.get_figheight()),
                      facecolor=BG)
    ax = disp.add_axes([0, 0, 1, 1])
    ax.imshow(img)
    ax.axis('off')
    disp.canvas.manager.set_window_title(label)
    plt.show(block=False)

def chart_picker(all_figs, df, title):
    """
    Interactive loop.  User picks ONE chart at a time (or multiple with commas).
    Each selection opens exactly that window and nothing else.
    'e' → export opened charts to Excel then return.
    'q' → discard and return.
    """
    setup_style()
    opened = []          # list of (fig, label) the user has viewed

    while True:
        print()
        cprint('  ┌─ CHART PICKER ─────────────────────────────────────────────┐', 'purple')
        for i, (ctype, label, _) in enumerate(all_figs, 1):
            tag     = f"{CHART_TAG_COLOR.get(ctype,'')}[{ctype:<7}]\033[0m"
            already = '\033[90m ✓\033[0m' if any(lb == label for _, lb in opened) else '  '
            print(f'  │  [{i}]{already} {tag} {label}')
        cprint('  ├─────────────────────────────────────────────────────────────┤', 'purple')
        cprint('  │  [a]  open All charts                                        │', 'white')
        cprint('  │  [e]  Export to Excel  (only charts you opened)              │', 'green')
        cprint('  │  [q]  Back to main menu                                      │', 'gray')
        cprint('  └─────────────────────────────────────────────────────────────┘', 'purple')
        print()
        cprint('  Tip: comma-separate numbers to open several at once, e.g. 1,3', 'gray')
        print()

        raw = prompt('  → Your choice: ').strip().lower()

        # ── quit ──────────────────────────────────────────────────────
        if raw == 'q':
            plt.close('all')
            return

        # ── export ────────────────────────────────────────────────────
        if raw == 'e':
            if not opened:
                cprint('  [!] You haven\'t opened any charts yet.', 'red')
                continue
            _do_export(opened, df, title)
            plt.close('all')
            return

        # ── open all ─────────────────────────────────────────────────
        if raw == 'a':
            for ctype, label, fig in all_figs:
                _show_via_savefig(fig, label)
                if not any(lb == label for _, lb in opened):
                    opened.append((fig, label))
            cprint(f'\n  ✓  All {len(all_figs)} charts opened.\n', 'green')
            continue

        # ── parse number(s) ───────────────────────────────────────────
        try:
            picks = [int(x.strip()) for x in raw.replace(' ', ',').split(',') if x.strip()]
        except ValueError:
            cprint('  [!] Enter a number, "a", "e", or "q".', 'red')
            continue

        invalid = [p for p in picks if p < 1 or p > len(all_figs)]
        if invalid:
            cprint(f'  [!] Invalid numbers: {invalid}. Valid: 1–{len(all_figs)}.', 'red')
            continue

        for pick in picks:
            ctype, label, fig = all_figs[pick - 1]
            _show_via_savefig(fig, label)
            if not any(lb == label for _, lb in opened):
                opened.append((fig, label))
            cprint(f'  ✓  Opened → {label}', 'green')

# ╔══════════════════════════════════════════════════════════════════╗
# ║  EXCEL EXPORT ENGINE                                            ║
# ╚══════════════════════════════════════════════════════════════════╝

def _fig_png_bytes(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=130, bbox_inches='tight',
                facecolor=fig.get_facecolor())
    buf.seek(0)
    return buf

def _thin():
    s = Side(style='thin', color='30363D')
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell, text, bg='1F2D3D', fg='E6EDF3'):
    cell.value = text
    cell.font      = Font(bold=True, color=fg, name='Consolas', size=10)
    cell.fill      = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = _thin()

def _dat(cell, val, even=True):
    cell.value     = val
    cell.fill      = PatternFill('solid', start_color='161B22' if even else '1C2230')
    cell.font      = Font(color='C9D1D9', name='Consolas', size=9)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = _thin()

def _write_df(ws, df, sr=1, sc=1):
    for ci, col in enumerate(df.columns, sc):
        _hdr(ws.cell(sr, ci), str(col))
        ws.column_dimensions[get_column_letter(ci)].width = max(14, len(str(col)) + 4)
    for ri, row in enumerate(df.itertuples(index=False), sr + 1):
        for ci, val in enumerate(row, sc):
            _dat(ws.cell(ri, ci), val, ri % 2 == 0)
    ws.row_dimensions[sr].height = 22

def _do_export(opened_figs, df, title):
    """Build .xlsx with Data, Charts (only opened ones), README sheets."""
    default = title.lower().replace(' ', '_') + '.xlsx'
    fname   = prompt(f'\n  → Filename [{default}]: ').strip()
    if not fname:
        fname = default
    if not fname.endswith('.xlsx'):
        fname += '.xlsx'

    cprint('\n  ⟳  Building Excel workbook...', 'cyan')
    wb = Workbook()

    # ── Sheet 1 : Data ───────────────────────────────────────────────
    ws_d = wb.active
    ws_d.title = 'Data'
    ws_d.sheet_view.showGridLines = False
    ws_d.sheet_properties.tabColor = '58A6FF'

    ws_d.merge_cells('A1:H1')
    ws_d['A1'].value     = f'  {title}  ·  dataviz.py v2.1'
    ws_d['A1'].font      = Font(bold=True, color='58A6FF', name='Consolas', size=13)
    ws_d['A1'].fill      = PatternFill('solid', start_color='0D1117')
    ws_d['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws_d.row_dimensions[1].height = 30

    ws_d.merge_cells('A2:H2')
    ws_d['A2'].value = (f'  Exported: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}'
                        f'  ·  Rows: {len(df)}  ·  Cols: {len(df.columns)}'
                        f'  ·  Charts embedded: {len(opened_figs)}')
    ws_d['A2'].font = Font(color='6E7681', name='Consolas', size=9, italic=True)
    ws_d['A2'].fill = PatternFill('solid', start_color='0D1117')
    ws_d.row_dimensions[2].height = 18

    _write_df(ws_d, df, sr=4, sc=1)
    ws_d.freeze_panes = 'A5'

    num_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    if num_cols:
        stats_df = df[num_cols].describe().round(2).reset_index()
        stats_df.rename(columns={'index': 'Stat'}, inplace=True)
        sr = len(df) + 7
        ws_d.merge_cells(f'A{sr}:H{sr}')
        ws_d[f'A{sr}'].value = '  STATISTICS  (df.describe())'
        ws_d[f'A{sr}'].font  = Font(bold=True, color='3FB950', name='Consolas', size=10)
        ws_d[f'A{sr}'].fill  = PatternFill('solid', start_color='0D1117')
        ws_d.row_dimensions[sr].height = 22
        _write_df(ws_d, stats_df, sr=sr + 1, sc=1)

    # ── Sheet 2 : Charts (only what the user opened) ─────────────────
    ws_c = wb.create_sheet('Charts')
    ws_c.sheet_view.showGridLines = False
    ws_c.sheet_properties.tabColor = 'D2A8FF'

    for r in range(1, 120):
        for c in range(1, 30):
            ws_c.cell(r, c).fill = PatternFill('solid', start_color='0D1117')

    ws_c.merge_cells('A1:P1')
    ws_c['A1'].value     = f'  {title}  ·  Charts ({len(opened_figs)} selected)'
    ws_c['A1'].font      = Font(bold=True, color='58A6FF', name='Consolas', size=13)
    ws_c['A1'].fill      = PatternFill('solid', start_color='0D1117')
    ws_c['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws_c.row_dimensions[1].height = 30

    ws_c.merge_cells('A2:P2')
    ws_c['A2'].value = f'  matplotlib v{matplotlib.__version__}  ·  Only user-selected charts are included'
    ws_c['A2'].font  = Font(color='6E7681', name='Consolas', size=9, italic=True)
    ws_c['A2'].fill  = PatternFill('solid', start_color='0D1117')
    ws_c.row_dimensions[2].height = 18

    COL_POS   = [2, 12]
    ROW_START = 4
    ROW_STEP  = 32

    for idx, (fig, label) in enumerate(opened_figs):
        buf  = _fig_png_bytes(fig)
        img  = XLImage(buf)
        img.width  = 620
        img.height = 390
        col_idx = COL_POS[idx % 2]
        row_pos = ROW_START + (idx // 2) * ROW_STEP
        col_ltr = get_column_letter(col_idx)

        lc = ws_c.cell(row=row_pos - 1, column=col_idx)
        lc.value = f'  {label}'
        lc.font  = Font(bold=True, color='8B949E', name='Consolas', size=9)
        lc.fill  = PatternFill('solid', start_color='0D1117')

        ws_c.add_image(img, f'{col_ltr}{row_pos}')

    # ── Sheet 3 : README ─────────────────────────────────────────────
    ws_r = wb.create_sheet('README')
    ws_r.sheet_view.showGridLines = False
    ws_r.sheet_properties.tabColor = '3FB950'

    readme_lines = [
        ('', ''),
        ('  DataViz Terminal Dashboard  v2.1', '58A6FF'),
        ('', ''),
        ('  EXPORT SUMMARY', '3FB950'),
        ('  ─────────────────────────────────────────────', '30363D'),
        (f'  Title      :  {title}', 'C9D1D9'),
        (f'  Exported   :  {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', 'C9D1D9'),
        (f'  Data rows  :  {len(df)}', 'C9D1D9'),
        (f'  Charts     :  {len(opened_figs)} (user-selected)', 'C9D1D9'),
        ('', ''),
        ('  CHARTS INCLUDED', '3FB950'),
        ('  ─────────────────────────────────────────────', '30363D'),
    ] + [(f'  {i+1}.  {lbl}', 'C9D1D9') for i, (_, lbl) in enumerate(opened_figs)] + [
        ('', ''),
        ('  SHEETS', '3FB950'),
        ('  ─────────────────────────────────────────────', '30363D'),
        ('  Data    :  Raw data + df.describe() statistics', 'C9D1D9'),
        ('  Charts  :  Selected matplotlib charts as images', 'C9D1D9'),
        ('  README  :  This page', 'C9D1D9'),
        ('', ''),
        ('  LIBRARIES', '3FB950'),
        ('  ─────────────────────────────────────────────', '30363D'),
        ('  matplotlib · pandas · numpy · openpyxl', 'C9D1D9'),
    ]

    for ri, (text, color) in enumerate(readme_lines, 1):
        cell = ws_r.cell(ri, 1, text)
        is_title = 'DataViz' in text
        is_hdr   = any(k in text for k in ['EXPORT', 'SHEETS', 'CHARTS', 'LIBRARIES'])
        cell.font = Font(color=color or '0D1117', name='Consolas',
                         size=13 if is_title else 10, bold=is_title or is_hdr)
        cell.fill = PatternFill('solid', start_color='0D1117')
        ws_r.row_dimensions[ri].height = 20

    ws_r.column_dimensions['A'].width = 60

    wb.save(fname)
    cprint(f'\n  ✓  Saved → {os.path.abspath(fname)}', 'green', bold=True)
    cprint(f'     Sheets : Data · Charts ({len(opened_figs)} charts) · README', 'gray')
    print()
    input('  Press Enter to return to menu...')

# ╔══════════════════════════════════════════════════════════════════╗
# ║  DATASET BUILDERS                                               ║
# ╚══════════════════════════════════════════════════════════════════╝

def build_students():
    np.random.seed(42)
    names    = ['Arjun S.','Priya M.','Rahul K.','Sneha T.','Dev R.',
                'Meera J.','Kunal P.','Ananya V.','Rohan B.','Pooja N.',
                'Vikram S.','Divya K.','Amit T.','Riya M.','Saurabh P.']
    subjects = ['Math','Science','English','History','Art']
    df = pd.DataFrame({s: np.random.randint(48, 99, len(names)) for s in subjects})
    df.insert(0, 'Student', names)
    df['Average'] = df[subjects].mean(axis=1).round(1)
    df['Grade']   = df['Average'].apply(
        lambda x: 'A' if x >= 85 else ('B' if x >= 70 else ('C' if x >= 55 else 'D')))
    months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug']
    trends = {'Math':    [72,75,70,78,82,80,85,88],
              'Science': [65,68,71,74,70,76,79,83],
              'English': [80,78,82,79,83,85,81,87]}
    gc = df['Grade'].value_counts()
    return {'type':'students','df':df,'subjects':subjects,'months':months,
            'trends':trends,'grade_labels':list(gc.index),
            'grade_vals':list(gc.values),'title':'Student Performance Dashboard'}

def build_sales():
    months  = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug']
    revenue = {'Electronics':    [320,290,380,410,350,430,480,510],
               'Clothing':       [180,210,170,195,230,215,240,260],
               'Home & Kitchen': [90,85,110,120,100,130,115,140]}
    regions  = ['North','South','East','West','Central','Online']
    reg_sale = [820,650,490,710,380,1050]
    cat_sh   = {'Electronics':42,'Clothing':28,'Home & Kitchen':15,'Books':9,'Sports':6}
    products = ['Smartphone','Earbuds','Kurta Set','Smart Watch',
                'Pressure Cooker','Yoga Mat','Backpack','Novel']
    units    = [482,316,580,201,340,410,228,612]
    prices   = [2000,1000,400,3000,500,300,500,120]
    df = pd.DataFrame({'Product':products,'Units':units,
                       'Price (Rs)':prices,
                       'Revenue (Rs)':[u*p for u,p in zip(units,prices)]})
    return {'type':'sales','df':df,'months':months,'revenue':revenue,
            'regions':regions,'reg_sales':reg_sale,'cat_share':cat_sh,
            'title':'Sales Analytics Dashboard'}

def build_figs_for(data):
    """Return list of (chart_type, label, fig). Figures closed immediately after creation."""
    if data['type'] == 'students':
        subjs = data['subjects']
        avgs  = [data['df'][s].mean() for s in subjs]
        return [
            ('line',    'Score Trends — Line Chart',
             make_line_fig(data['months'], data['trends'], 'Score Trends by Subject')),
            ('bar',     'Subject Averages — Bar Chart',
             make_bar_fig(subjs, avgs, 'Avg Score', 'Average Score per Subject')),
            ('pie',     'Grade Distribution — Pie Chart',
             make_pie_fig(data['grade_labels'], data['grade_vals'], 'Grade Distribution')),
            ('barh',    'Student Rankings — Barh Chart',
             make_barh_fig(list(data['df']['Student']),
                           list(data['df']['Average']),
                           'Average Score', 'Student Rankings')),
            ('scatter', 'Score Distribution — Scatter Plot',
             make_scatter_fig(list(range(1, len(data['df'])+1)),
                              list(data['df']['Average']),
                              'Student Index', 'Average Score', 'Score Distribution')),
        ]
    if data['type'] == 'sales':
        return [
            ('line',    'Revenue Trends — Line Chart',
             make_line_fig(data['months'], data['revenue'], 'Monthly Revenue by Category')),
            ('bar',     'Regional Sales — Bar Chart',
             make_bar_fig(data['regions'], data['reg_sales'], 'Revenue (K)', 'Sales by Region')),
            ('pie',     'Category Share — Pie Chart',
             make_pie_fig(list(data['cat_share'].keys()),
                          list(data['cat_share'].values()), 'Revenue by Category')),
            ('barh',    'Product Revenue — Barh Chart',
             make_barh_fig(list(data['df']['Product']),
                           list(data['df']['Revenue (Rs)']),
                           'Revenue (Rs)', 'Revenue by Product')),
            ('scatter', 'Units vs Revenue — Scatter Plot',
             make_scatter_fig(list(data['df']['Units']),
                              list(data['df']['Revenue (Rs)']),
                              'Units Sold', 'Revenue (Rs)', 'Units vs Revenue')),
        ]
    return []

# ╔══════════════════════════════════════════════════════════════════╗
# ║  MODES                                                          ║
# ╚══════════════════════════════════════════════════════════════════╝

PRELOADED = {
    '1': ('Student Performance', '15 students · 5 subjects · grades', build_students),
    '2': ('Sales Analytics',     'Revenue · regions · category share',  build_sales),
}

def preloaded_menu():
    header()
    cprint('  ┌─ PRE-LOADED DATASETS ──────────────────────────────────────┐', 'blue')
    for k, (name, desc, _) in PRELOADED.items():
        cprint(f'  │  [{k}]  {name:<24} {desc:<34}│', 'white')
    cprint('  │  [0]  ← Back                                               │', 'gray')
    cprint('  └─────────────────────────────────────────────────────────────┘', 'blue')
    print()

    while True:
        choice = prompt('  → Select dataset: ').strip()
        if choice == '0':
            return
        if choice in PRELOADED:
            name, _, builder = PRELOADED[choice]
            cprint(f'\n  ⟳  Loading "{name}"...', 'cyan')
            data = builder()
            print()
            cprint(f'  DATA PREVIEW  ·  {name}', 'purple', bold=True)
            print_df(data['df'])
            go = prompt('  → Continue to chart picker? [Y/n]: ').strip().lower()
            if go in ('', 'y', 'yes'):
                cprint('\n  ⟳  Building chart objects...', 'cyan')
                all_figs = build_figs_for(data)
                cprint(f'  ✓  {len(all_figs)} charts ready.\n', 'green')
                chart_picker(all_figs, data['df'], data['title'])
            return
        cprint('  [!] Enter 1, 2, or 0.', 'red')

def manual_entry():
    header()
    cprint('  ┌─ MANUAL DATA ENTRY MODE ──────────────────────────────────┐', 'cyan')
    cprint('  │  Type category names and values directly in the terminal  │', 'cyan')
    cprint('  └───────────────────────────────────────────────────────────┘', 'cyan')
    print()

    title = prompt('  → Dashboard title: ').strip() or 'Custom Dashboard'
    label = prompt('  → Y-axis label (e.g. Score, Amount): ').strip() or 'Value'
    print()

    while True:
        try:
            n = int(prompt('  → Number of categories (2-20): '))
            if 2 <= n <= 20:
                break
            cprint('  [!] Enter between 2 and 20.', 'red')
        except ValueError:
            cprint('  [!] Whole number please.', 'red')

    print()
    cprint(f'  Enter {n} entries:', 'yellow')
    divider()

    categories, values = [], []
    for i in range(1, n + 1):
        while True:
            name = prompt(f'  [{i}/{n}] Category : ').strip()
            if name:
                break
            cprint('       [!] Cannot be empty.', 'red')
        while True:
            try:
                val = float(prompt(f'  [{i}/{n}] {label:<12}: '))
                break
            except ValueError:
                cprint('       [!] Enter a number.', 'red')
        categories.append(name)
        values.append(val)
        cprint(f'       ✓  {name} = {val}', 'green')

    df = pd.DataFrame({'Category': categories, label: values})
    df['% Share'] = (df[label].abs() / df[label].abs().sum() * 100).round(1)

    print()
    divider()
    cprint(f'  PREVIEW  ·  {title}', 'purple', bold=True)
    print_df(df, max_rows=25)
    cprint(f'  Min: {min(values):.2f}  │  Max: {max(values):.2f}  │  '
           f'Mean: {sum(values)/len(values):.2f}  │  Total: {sum(values):.2f}', 'gray')
    divider()
    print()

    go = prompt('  → Continue to chart picker? [Y/n]: ').strip().lower()
    if go not in ('', 'y', 'yes'):
        cprint('\n  Aborted.\n', 'gray')
        return

    abs_vals = [abs(v) for v in values]
    cprint('\n  ⟳  Building chart objects...', 'cyan')
    all_figs = [
        ('bar',     f'{label} Comparison — Bar Chart',
         make_bar_fig(categories, values, label, f'{label} Comparison')),
        ('barh',    f'{label} Ranked — Barh Chart',
         make_barh_fig(categories, values, label, f'{label} Ranked')),
        ('pie',     f'Distribution — Pie Chart',
         make_pie_fig(categories, abs_vals, 'Share Distribution')),
        ('line',    f'{label} Trend — Line Chart',
         make_line_fig(categories, {label: values}, 'Trend by Category')),
        ('scatter', f'{label} Scatter Plot',
         make_scatter_fig(list(range(1, len(values)+1)), values,
                          'Index', label, f'{label} Distribution')),
    ]
    cprint(f'  ✓  {len(all_figs)} charts ready.\n', 'green')
    chart_picker(all_figs, df, title)

def csv_import():
    header()
    cprint('  ┌─ CSV IMPORT MODE ──────────────────────────────────────────┐', 'yellow')
    cprint('  │  Paste the full path to your .csv file                     │', 'yellow')
    cprint('  └─────────────────────────────────────────────────────────────┘', 'yellow')
    print()

    while True:
        raw = prompt('  → CSV file path: ').strip().strip('"').strip("'")
        if not raw:
            cprint('  [!] No path entered.', 'red')
            continue
        if not os.path.isfile(raw):
            cprint(f'  [!] Not found: {raw}', 'red')
            continue
        break

    cprint(f'\n  ⟳  Reading {os.path.basename(raw)} ...', 'cyan')
    try:
        df = pd.read_csv(raw)
    except Exception as e:
        cprint(f'  [ERROR] {e}', 'red')
        input('  Press Enter...'); return

    cprint(f'  ✓  Loaded {len(df)} rows × {len(df.columns)} columns', 'green')
    print()
    print_df(df)

    num_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    cat_cols = df.select_dtypes(exclude=[np.number]).columns.tolist()

    if not num_cols:
        cprint('  [!] No numeric columns — cannot plot.', 'red')
        input('  Press Enter...'); return

    cprint(f'  Numeric : {", ".join(num_cols)}', 'green')
    cprint(f'  Text    : {", ".join(cat_cols) if cat_cols else "none"}', 'gray')
    print()

    label_col = cat_cols[0] if cat_cols else None
    title     = os.path.splitext(os.path.basename(raw))[0].replace('_', ' ').title()

    go = prompt('  → Continue to chart picker? [Y/n]: ').strip().lower()
    if go not in ('', 'y', 'yes'):
        input('  Aborted. Press Enter...'); return

    cprint('\n  ⟳  Building chart objects...', 'cyan')
    all_figs = []

    for col in num_cols:
        vals = df[col].dropna().tolist()
        cats = (df[label_col].astype(str).tolist()
                if label_col else [str(i+1) for i in range(len(vals))])
        cats = cats[:len(vals)]

        all_figs.append(('bar',  f'{col} — Bar Chart',
            make_bar_fig(cats[:20], vals[:20], col, f'{col} by {label_col or "Index"}')))
        all_figs.append(('barh', f'{col} — Barh Chart',
            make_barh_fig(cats[:15], vals[:15], col, f'{col} Ranked')))
        all_figs.append(('line', f'{col} — Line Chart',
            make_line_fig([str(c)[:12] for c in cats[:20]], {col: vals[:20]}, f'{col} Trend')))

    if label_col and len(df) <= 12:
        all_figs.append(('pie', f'{num_cols[0]} — Pie Chart',
            make_pie_fig(df[label_col].astype(str).tolist(),
                         df[num_cols[0]].abs().tolist(), f'{num_cols[0]} Distribution')))

    if len(num_cols) >= 2:
        xv = df[num_cols[0]].dropna().tolist()
        yv = df[num_cols[1]].dropna().tolist()
        ml = min(len(xv), len(yv))
        all_figs.append(('scatter', f'{num_cols[0]} vs {num_cols[1]} — Scatter',
            make_scatter_fig(xv[:ml], yv[:ml], num_cols[0], num_cols[1],
                             f'{num_cols[0]} vs {num_cols[1]}')))

    cprint(f'  ✓  {len(all_figs)} charts ready.\n', 'green')
    chart_picker(all_figs, df, title)

# ╔══════════════════════════════════════════════════════════════════╗
# ║  MAIN MENU                                                      ║
# ╚══════════════════════════════════════════════════════════════════╝

def main_menu():
    while True:
        header()
        cprint('  ┌─ MAIN MENU ────────────────────────────────────────────────┐', 'green')
        cprint('  │                                                             │', 'green')
        cprint('  │   [1]  Pre-loaded datasets   Students / Sales              │', 'white')
        cprint('  │   [2]  Manual data entry     Type your own data            │', 'white')
        cprint('  │   [3]  Import CSV            Auto-detect & plot             │', 'white')
        cprint('  │   [q]  Quit                                                 │', 'gray')
        cprint('  │                                                             │', 'green')
        cprint('  └─────────────────────────────────────────────────────────────┘', 'green')
        print()
        cprint('  Charts : line · bar · barh · scatter · pie  (you pick which)', 'gray')
        cprint('  Export : Excel (.xlsx) — only selected charts are embedded', 'gray')
        print()

        choice = prompt('  → Select option: ').strip().lower()
        if   choice == '1': preloaded_menu()
        elif choice == '2': manual_entry()
        elif choice == '3': csv_import()
        elif choice in ('q', 'quit', 'exit'):
            print()
            cprint('  Goodbye! Happy plotting! 🐍', 'cyan', bold=True)
            print()
            sys.exit(0)
        else:
            cprint('\n  [!] Invalid. Enter 1, 2, 3, or q.\n', 'red')
            input('  Press Enter to continue...')

if __name__ == '__main__':
    try:
        main_menu()
    except Exception as e:
        print(f"\n[ERROR] {e}")
    input("\nPress Enter to exit...")